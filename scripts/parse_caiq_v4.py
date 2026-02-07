#!/usr/bin/env python3
"""
Parse CAIQ v4.1 Excel file to JSON format for TPRM Frameworks MCP.

This script reads the CAIQv4.1.0 sheet and generates a complete JSON dataset
with 283 questions following the project's data structure.
"""

import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional


# Column mappings (0-indexed)
COL_QUESTION_ID = 0  # A
COL_QUESTION = 1  # B
COL_CCM_CONTROL_ID = 8  # I
COL_CCM_CONTROL_SPEC = 9  # J
COL_CCM_CONTROL_TITLE = 10  # K
COL_CCM_DOMAIN_TITLE = 11  # L


def determine_answer_type(question_text: str) -> str:
    """Determine the expected answer type based on question text."""
    question_lower = question_text.lower().strip()

    # Check for yes/no questions
    yes_no_patterns = ['are ', 'is ', 'do ', 'does ', 'has ', 'have ', 'can ', 'will ', 'should ']
    for pattern in yes_no_patterns:
        if question_lower.startswith(pattern):
            return "yes_no"

    return "text"


def determine_risk_level(domain: str, control_title: str) -> str:
    """Determine risk level based on domain and control title."""
    domain_lower = domain.lower()
    title_lower = control_title.lower()

    # High-risk domains and keywords
    high_risk_keywords = [
        'cryptography', 'encryption', 'access control', 'authentication',
        'audit', 'data security', 'network security', 'vulnerability',
        'threat', 'incident', 'breach', 'compliance'
    ]

    for keyword in high_risk_keywords:
        if keyword in domain_lower or keyword in title_lower:
            return "high"

    # Medium risk for governance, HR, physical security
    medium_risk_keywords = ['governance', 'human resource', 'physical', 'asset']
    for keyword in medium_risk_keywords:
        if keyword in domain_lower or keyword in title_lower:
            return "medium"

    # Default to medium
    return "medium"


def determine_weight(risk_level: str) -> int:
    """Determine question weight based on risk level."""
    if risk_level == "high":
        return 8
    elif risk_level == "medium":
        return 5
    else:
        return 3


def generate_evaluation_rubric(question_type: str, control_spec: str) -> Dict[str, Any]:
    """Generate evaluation rubric based on question type and control specification."""
    if question_type == "yes_no":
        # Extract keywords from control specification
        keywords = []
        keyword_candidates = ['policy', 'procedure', 'documented', 'implemented',
                            'approved', 'reviewed', 'maintained', 'established',
                            'communicated', 'evaluated', 'monitored']

        spec_lower = control_spec.lower()
        for candidate in keyword_candidates:
            if candidate in spec_lower:
                keywords.append(candidate)

        return {
            "acceptable": ["yes", "implemented", "documented", "in place", "established"],
            "partially_acceptable": ["in progress", "planned", "partially", "working on"],
            "unacceptable": ["no", "not implemented", "not established", "none", "n/a"],
            "required_keywords": keywords[:3] if keywords else ["policy", "procedure"]
        }
    else:
        # For text questions, require detailed responses
        return {
            "acceptable": ["detailed", "comprehensive", "specific"],
            "partially_acceptable": ["some", "limited", "basic"],
            "unacceptable": ["none", "not applicable", "n/a"],
            "required_keywords": []
        }


def parse_caiq_excel(excel_path: str) -> Dict[str, Any]:
    """Parse CAIQ v4.1 Excel file into JSON structure."""

    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['CAIQv4.1.0']

    questions = []

    # Track last CCM control for inheritance
    last_ccm_control_id = None
    last_ccm_control_spec = None
    last_ccm_control_title = None
    last_ccm_domain_title = None

    # Statistics
    stats = {
        'total_questions': 0,
        'domains': {},
        'question_types': {'yes_no': 0, 'text': 0},
        'risk_levels': {'high': 0, 'medium': 0, 'low': 0}
    }

    print("Parsing questions...")

    # Start from row 3 (data starts after header row 2)
    for row_idx in range(3, ws.max_row + 1):
        row = ws[row_idx]

        # Extract values
        question_id = row[COL_QUESTION_ID].value
        question_text = row[COL_QUESTION].value
        ccm_control_id = row[COL_CCM_CONTROL_ID].value
        ccm_control_spec = row[COL_CCM_CONTROL_SPEC].value
        ccm_control_title = row[COL_CCM_CONTROL_TITLE].value
        ccm_domain_title = row[COL_CCM_DOMAIN_TITLE].value

        # Skip rows without question ID
        if not question_id:
            continue

        # Skip rows without question text
        if not question_text:
            continue

        # Handle CCM field inheritance
        if ccm_control_id:
            last_ccm_control_id = ccm_control_id
        if ccm_control_spec:
            last_ccm_control_spec = ccm_control_spec
        if ccm_control_title:
            last_ccm_control_title = ccm_control_title
        if ccm_domain_title:
            last_ccm_domain_title = ccm_domain_title

        # Use last known values if current is None
        current_ccm_id = ccm_control_id or last_ccm_control_id
        current_ccm_spec = ccm_control_spec or last_ccm_control_spec
        current_ccm_title = ccm_control_title or last_ccm_control_title
        current_domain = ccm_domain_title or last_ccm_domain_title

        # Clean up text (remove extra whitespace and newlines)
        question_text = ' '.join(question_text.split())
        current_ccm_spec = ' '.join(current_ccm_spec.split()) if current_ccm_spec else ""

        # Determine question attributes
        answer_type = determine_answer_type(question_text)
        risk_level = determine_risk_level(current_domain, current_ccm_title)
        weight = determine_weight(risk_level)

        # Generate evaluation rubric
        rubric = generate_evaluation_rubric(answer_type, current_ccm_spec)

        # Build question object
        question = {
            "id": question_id.strip(),
            "category": current_domain.strip(),
            "subcategory": current_ccm_title.strip(),
            "question_text": question_text.strip(),
            "description": current_ccm_spec.strip(),
            "expected_answer_type": answer_type,
            "is_required": True,
            "weight": weight,
            "regulatory_mappings": ["ISO 27001:2022", "SOC 2", "NIST CSF"],
            "scf_control_mappings": [],  # Will be filled by Agent 2
            "risk_if_inadequate": risk_level,
            "ccm_control_id": current_ccm_id.strip() if current_ccm_id else "",
            "evaluation_rubric": rubric
        }

        questions.append(question)

        # Update statistics
        stats['total_questions'] += 1
        stats['domains'][current_domain] = stats['domains'].get(current_domain, 0) + 1
        stats['question_types'][answer_type] += 1
        stats['risk_levels'][risk_level] += 1

        if stats['total_questions'] % 50 == 0:
            print(f"  Processed {stats['total_questions']} questions...")

    # Get unique domains for metadata
    domains = sorted(list(set(q['category'] for q in questions)))

    # Build final JSON structure (matching expected format)
    result = {
        "metadata": {
            "name": "CSA CAIQ v4.1.0",
            "version": "4.1.0",
            "total_questions": len(questions),
            "status": "production",
            "description": "Complete Cloud Security Alliance Consensus Assessment Initiative Questionnaire v4.1.0 - 283 questions mapped to Cloud Controls Matrix (CCM) v4",
            "categories": domains,
            "estimated_completion_time": "20-30 hours"
        },
        "questions": questions
    }

    print(f"\nParsing complete!")
    print(f"Total questions: {stats['total_questions']}")
    print(f"\nQuestions by domain:")
    for domain, count in sorted(stats['domains'].items()):
        print(f"  {domain}: {count}")
    print(f"\nQuestion types:")
    print(f"  Yes/No: {stats['question_types']['yes_no']}")
    print(f"  Text: {stats['question_types']['text']}")
    print(f"\nRisk levels:")
    print(f"  High: {stats['risk_levels']['high']}")
    print(f"  Medium: {stats['risk_levels']['medium']}")
    print(f"  Low: {stats['risk_levels']['low']}")

    return result, stats


def main():
    """Main execution function."""

    # File paths
    excel_path = "/tmp/CCM+CAIQv4.1 Bundle/CAIQv4.1.0-star_security_questionnaire-generated_at_2026_01_13.xlsx"
    output_path = "/Users/jeffreyvonrotz/Projects/TPRM-Frameworks-mcp/src/tprm_frameworks_mcp/data/caiq_v4_full.json"

    # Parse Excel
    result, stats = parse_caiq_excel(excel_path)

    # Write JSON
    print(f"\nWriting to {output_path}...")
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"✅ Successfully created {output_path}")
    print(f"✅ File size: {output_file.stat().st_size / 1024:.1f} KB")

    # Validation
    print("\nValidating output...")
    with open(output_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Validate metadata structure
    assert 'metadata' in data, "Missing metadata section"
    assert 'questions' in data, "Missing questions section"
    assert data['metadata']['total_questions'] == stats['total_questions'], "Question count mismatch in metadata"
    assert len(data['questions']) == stats['total_questions'], "Question count mismatch in questions array"

    # Check first question has all required fields
    first_q = data['questions'][0]
    required_fields = ['id', 'category', 'subcategory', 'question_text', 'description',
                      'expected_answer_type', 'is_required', 'weight', 'regulatory_mappings',
                      'scf_control_mappings', 'risk_if_inadequate', 'ccm_control_id',
                      'evaluation_rubric']

    for field in required_fields:
        assert field in first_q, f"Missing required field: {field}"

    print("✅ Validation passed!")
    print(f"\n🎉 CAIQ v4.1 parsing complete: {stats['total_questions']} questions")


if __name__ == "__main__":
    main()
